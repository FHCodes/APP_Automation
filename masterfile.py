import pandas as pd
import os
from alerta import msg_alerta_alert, msg_alerta_erro, msg_alerta_sucesso
import zipfile
import openpyxl
import shutil

def cria_masterfiles(path_name_xlsm,inventory_name,status,caminho_master):
    path_name = criando_pasta(path_name_xlsm)

    if len(inventory_name) > len(set(inventory_name)):
        deleta_pasta(path_name)
        msg_alerta_erro("Ocorreu um erro ao gerar as masterfiles!","Verifique se na coluna Inventory name da aba 3. Data Sources há tabelas com o mesmo nome!")
        raise SystemExit
    
    Data_Sources_list, table_name_dic,table_name_list, Data_Sources_Attr_list, pk_list, Data_Sources_Map_list = reducao_dados(path_name_xlsm,inventory_name)

    if not Data_Sources_list and len(inventory_name) > 0 :
        deleta_pasta(path_name)
        msg_alerta_erro('Ocorreu um erro ao gerar as masterfiles!','Nenhum Inventory Name encontrado no pack')
        
    elif len(table_name_list) <= len(inventory_name) and inventory_name[0] != '':
        diff = set(table_name_list).symmetric_difference(set(inventory_name))
        if diff:
            for dif in list(diff):
                msg_alerta_alert(dif,'Inventory Name não encontrado no pack')
        try:
            if status:
                mtf_prt1_oracle(inventory_name,Data_Sources_list, path_name)
            else:
                mtf_prt1_postgre(inventory_name,Data_Sources_list, path_name)

            if status:
                mtf_prt2_oracle(inventory_name, Data_Sources_Attr_list,path_name)
            else:
                mtf_prt2_postgre(inventory_name, Data_Sources_Attr_list,path_name)

        except BaseException as err:
            msg_alerta_erro('Ocorreu um erro ao gerar as masterfiles!',f'{err}')
            deleta_pasta(path_name)

        try:   
               
            if not caminho_master:
                mtf_prt3_sem_caminho(inventory_name, Data_Sources_Map_list, table_name_dic, path_name)
            else:
                mtf_prt3_com_caminho(inventory_name, Data_Sources_Map_list, table_name_dic, path_name,caminho_master)
        except:
            deleta_pasta(path_name)
            msg_alerta_erro('Ocorreu um erro ao tentar gerar as masterfiles!','Favor preencher as colunas "Enrichment Attribute Name", "DBN0 Attribute Name" e "AdHoc Join Type" presentes na aba 3. Data Source Map.".')
        
        acx(inventory_name, Data_Sources_list, pk_list, path_name)
        
        msg_alerta_sucesso('SUCESSO','MASTERFILES E ACX CRIADOS')
        
    else:
        msg_alerta_erro("Ocorreu um erro ao gerar as masterfiles!","Verifique se existem valores repetidos na coluna Inventory Name")
        raise SystemExit
                               
def criando_pasta(path_name_xlsm):
    temp = path_name_xlsm.split('/')
    nome_pack = temp.pop().split('.')
    nome_pack.pop()
    nome_pack = '.'.join(nome_pack)
    temp = '/'.join(temp)
    path_name = temp + '/MASTERFILES' +'_' + nome_pack

    if not os.path.exists(path_name):
        os.makedirs(path_name)
    
    return path_name

def deleta_pasta(path_name):
    shutil.rmtree(path_name)


def reducao_dados(path_name_xlsm,inventory_name):

    path_name = criando_pasta(path_name_xlsm)
    
    try:
        book = openpyxl.load_workbook(path_name_xlsm)
        Data_Sources = book['3. Data Sources']
        Data_Sources_Attr = book['3. Data Sources Attr & Count']
        Data_Sources_Map = book['3. Data Sources Map']
    
    except Exception as e:
        msg_alerta_erro('Ocorreu um erro ao gerar as masterfiles!','Limpe formatos e filtros da planilha e verifique se o nome das abas 3. estão corretos.')
        deleta_pasta(path_name)
    
    for rows in Data_Sources.iter_rows(min_row=4,max_row=4):
        
        if rows[1].value != "Inventory Name":
            msg_alerta_erro("Ocorreu um erro ao gerar as masterfiles!","O cabeçalho deve ficar na 4 linha")
            raise SystemExit

        for i, row in enumerate(rows):
            if row.value == 'Inventory Name':  # _0
                col_Inv_Name = i
            if row.value == 'Table Name':      # _1
                col_Tab_Name = i
            if row.value == 'Schema':          # Schema
                col_Schema = i
            if row.value == 'Description':     # Description
                col_Desc = i

    for rows in Data_Sources_Attr.iter_rows(min_row=4,max_row=4):
        
        if rows[1].value != "Source Name":
            msg_alerta_erro("Ocorreu um erro ao gerar as masterfiles!","O cabeçalho deve ficar na 4 linha")
            raise SystemExit
        
        for i, row in enumerate(rows):
            if row.value == 'Source Name':                          # _0
                col_SN = i
            if row.value == 'Attribute/Counter Name':               # _1
                col_ACN = i
            if row.value == 'Attribute/Counter Physical Name':      # _2
                col_ACPN = i
            if row.value == 'Data Type':                            # _3
                col_DT = i
            if row.value == 'Mediation Type':                       # _4
                col_MT = i
            if row.value == 'Metrics Attribute Type':               # _5
                col_MAT = i
            if row.value == 'Description':                          # Description
                col_D = i

    for rows in Data_Sources_Map.iter_rows(min_row=4,max_row=4):
        
        if rows[1].value.strip() != "Enrichment Table Name":
            msg_alerta_erro("Ocorreu um erro ao gerar as masterfiles!","O cabeçalho deve ficar na 4 linha")
            raise SystemExit
        
        for i, row in enumerate(rows):
            if row.value =='Enrichment Table Name':        # _0
                col_ETN = i
            if row.value =='Enrichment Attribute Name':    # _1
                col_EAN = i
            if row.value =='DBNO Table Name':              # _2
                col_DTN = i
            if row.value =='DBN0 Attribute Name':          # _3
                col_DAN = i
            if row.value =='AdHoc Join Type':              # _4
                col_AJT = i  

    Data_Sources = pd.read_excel(path_name_xlsm, sheet_name='3. Data Sources',usecols=[col_Inv_Name,col_Tab_Name,col_Schema,col_Desc],skiprows=3) 
    Data_Sources_list = []
    table_name_dic = {}
    table_name_list = []

    Data_Sources_Attr = pd.read_excel(path_name_xlsm, sheet_name='3. Data Sources Attr & Count',usecols=[col_SN,col_ACN,col_ACPN,col_DT,col_MT,col_MAT,col_D],skiprows=3) 
    Data_Sources_Attr.dropna(axis=0) # Exclui linhas que tenham Nan
    Data_Sources_Attr_list = []
    pk_list = []

    Data_Sources_Map = pd.read_excel(path_name_xlsm, sheet_name='3. Data Sources Map',usecols=[col_ETN, col_EAN, col_DTN, col_DAN, col_AJT],skiprows=3)  
    Data_Sources_Map_list = []
    
    for i in inventory_name:
    
      #Data_Sources
      for row in Data_Sources.itertuples(index=False):
          if row._0 == i:
              table_name_dic[row._0] = row._1
              table_name_list.append(row._0)

              if row.Description != None and "'" in str(row.Description):
                  desc = row.Description.split("'")
                  desc = "".join(desc)
                  Data_Sources_list.append((row._0,row._1,row.Schema.split("_")[-1].upper(),desc, row.Schema))
              else:
                  Data_Sources_list.append((row._0,row._1,row.Schema.split("_")[-1].upper(),row.Description, row.Schema))
      
    #Data_Sources_Attr
   
      for row in Data_Sources_Attr.itertuples(index=False):
          if row._0 == i:
              if "'" in str(row.Description): #Isso foi umas das coisas mais sem sentido que eu já vi, entender uma string como float
                  desc1 = row.Description.split("'")
                  desc1 = "".join(desc1)
                  Data_Sources_Attr_list.append((row._0,row._1,row._2,row._3,row._4,row._5,desc1))
              else:
                  Data_Sources_Attr_list.append((row._0,row._1,row._2,row._3,row._4,row._5,row.Description))       
            
              try:
                if row._4.upper() == 'PK':
                    pk_list.append((row._0,row._2,row._4))
              except:
                deleta_pasta(path_name)
                msg_alerta_erro('Ocorreu um erro ao gerar as masterfiles!','"Coluna Mediation Type da aba 3. Data Sources Attr & Count não pode estar vazia"')

      #Data_Sources_Map   
      for row in Data_Sources_Map.itertuples(index=False):
          if i == row._2:
            Data_Sources_Map_list.append((row._0,row._1, row._2,row._3,row._4))
    return Data_Sources_list, table_name_dic, table_name_list, Data_Sources_Attr_list, pk_list, Data_Sources_Map_list

def mtf_prt1_postgre(inventory_name,Data_Sources_list, path_name):   
    #Masterfiles - parte1
    
    for i in inventory_name:
            for ds in Data_Sources_list:
                if ds[2] != 'ENRICH' and ds[0] == i:
                    with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:            
                        arquivo.write(f"FILENAME={ds[1]}, SUFFIX=SQLPSTGR, REMARKS='{ds[3]}', $\n")
                        arquivo.write(f"  SEGMENT={ds[1]}, SEGTYPE=S0, $\n")
                        arquivo.write("FIELDNAME=SEQ_NUMBER, ALIAS=seq_number, USAGE=P21, ACTUAL=P11, $\n")
                        arquivo.write("FIELDNAME=INSERT_DATE, ALIAS=insert_date, USAGE=HYYMDs, ACTUAL=HYYMDs, $\n")
                        arquivo.write("FIELDNAME=SOURCE_ID, ALIAS=source_id, USAGE=A255V, ACTUAL=A255V, $\n")
                        arquivo.write("FIELDNAME=CONTENT_ID, ALIAS=content_id, USAGE=A256V, ACTUAL=A256V, $\n")
                
                elif ds[2] == 'ENRICH' and ds[0] == i:
                    with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:
                        arquivo.write(f"FILENAME={ds[1]}, SUFFIX=SQLPSTGR,\n")
                        arquivo.write(f"  SEGMENT={ds[1]}, SEGTYPE=S0, $\n")

def mtf_prt1_oracle(inventory_name,Data_Sources_list, path_name):   
    #Masterfiles - parte1
    
    for i in inventory_name:
            for ds in Data_Sources_list:
                if ds[2] != 'ENRICH' and ds[0] == i:
                    with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:            
                        arquivo.write(f"FILENAME={ds[1]}, SUFFIX=SQLORA, REMARKS='{ds[3]}', $\n")
                        arquivo.write(f"  SEGMENT={ds[1]}, SEGTYPE=S0, $\n")
                        arquivo.write("FIELDNAME=SEQ_NUMBER, ALIAS=SEQ_NUMBER, USAGE=P21, ACTUAL=P11, $\n")
                        arquivo.write("FIELDNAME=INSERT_DATE, ALIAS=INSERT_DATE, USAGE=HYYMDs, ACTUAL=HYYMDs, $\n")
                       # arquivo.write("FIELDNAME=SOURCE_ID, ALIAS=SOURCE_ID, USAGE=A255V, ACTUAL=A255V, $\n")
                       # arquivo.write("FIELDNAME=CONTENT_ID, ALIAS=CONTENT_ID, USAGE=A256V, ACTUAL=A256V, $\n")
                
                elif ds[2] == 'ENRICH' and ds[0] == i:
                    with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:
                        arquivo.write(f"FILENAME={ds[1]}, SUFFIX=SQLPSTGR,\n")
                        arquivo.write(f"  SEGMENT={ds[1]}, SEGTYPE=S0, $\n")
        
def mtf_prt2_postgre(inventory_name, Data_Sources_Attr_list,path_name):
    #Masterfiles - parte2
    for i in inventory_name:
        for dsa in Data_Sources_Attr_list:
            if dsa[0] == i:
                with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:
                    if dsa[3] == "TIMESTAMP(3)" and dsa[5] != "Constant":
                        arquivo.write(f"FIELDNAME={dsa[1].upper()}, ALIAS={dsa[2].lower()}, TITLE='{dsa[1]}', DESCRIPTION='{dsa[6]}',USAGE=HYYMDs, ACTUAL=HYYMDs,\n    MISSING=ON, $\n")
                    elif dsa[3][0:7] == "VARCHAR" and dsa[5] != "Constant":
                        arquivo.write(f"FIELDNAME={dsa[1].upper()}, ALIAS={dsa[2].lower()}, TITLE='{dsa[1]}', DESCRIPTION='{dsa[6]}',USAGE=A255V, ACTUAL=A255V,\n    MISSING=ON, $\n")
                    elif dsa[3] == "NUMBER" and dsa[5] != "Constant":
                        arquivo.write(f"FIELDNAME={dsa[1].upper()}, ALIAS={dsa[2].lower()}, TITLE='{dsa[1]}', DESCRIPTION='{dsa[6]}',USAGE=D20.2, ACTUAL=D8,\n    MISSING=ON, $\n")

def mtf_prt2_oracle(inventory_name, Data_Sources_Attr_list,path_name):
    #Masterfiles - parte2
    for i in inventory_name:
        for dsa in Data_Sources_Attr_list:
            if dsa[0] == i:
                with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:
                    if dsa[3] == "TIMESTAMP(3)" and dsa[5] != "Constant":
                        arquivo.write(f"FIELDNAME={dsa[1].upper()}, ALIAS={dsa[2].upper()}, TITLE='{dsa[1]}', DESCRIPTION='{dsa[6]}',USAGE=HYYMDs, ACTUAL=HYYMDs,\n    MISSING=ON, $\n")
                    elif dsa[3][0:7] == "VARCHAR" and dsa[5] != "Constant":
                        arquivo.write(f"FIELDNAME={dsa[1].upper()}, ALIAS={dsa[2].upper()}, TITLE='{dsa[1]}', DESCRIPTION='{dsa[6]}',USAGE=A255V, ACTUAL=A255V,\n    MISSING=ON, $\n")
                    elif dsa[3] == "NUMBER" and dsa[5] != "Constant":
                        arquivo.write(f"FIELDNAME={dsa[1].upper()}, ALIAS={dsa[2].upper()}, TITLE='{dsa[1]}', DESCRIPTION='{dsa[6]}',USAGE=D20.2, ACTUAL=D8,\n    MISSING=ON, $\n")                  
                    
def mtf_prt3_com_caminho(inventory_name,Data_Sources_Map_list,table_name_dic, path_name, caminho_master):                
   #Masterfiles - parte3
    dsm = Data_Sources_Map_list[:]
    
    for i in inventory_name:
        lista_ETN = []
        lista_res = []
        #Gera a lista de Enrichment Table Name
        for t in range(len(dsm)):
            if dsm[t][2] == i:
                lista_ETN.append(dsm[t][0])

    #Comando Set está mudando a ordem
        lista_ETN_uni = list(set(lista_ETN))


        for k in lista_ETN_uni:
            lista_res.append((k,lista_ETN.count(k)))
        
        for h in range(len(lista_res)):
          
            for j in range(len(dsm)):
                
                if dsm[j][2] == i and lista_res[h][0] == dsm[j][0]:
                    try:
                        if lista_res[h][1] == 1:
                            with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:
                                arquivo.write(f"""SEGMENT={dsm[j][0]}, SEGTYPE=KU,PARENT={table_name_dic[dsm[j][2]]}, CRFILE={caminho_master}/{dsm[j][0]}, CRINCLUDE=ALL , CRJOINTYPE={dsm[j][4].replace(' ', '_').upper()}, JOIN_WHERE={table_name_dic[dsm[j][2]]}.{dsm[j][3].upper()} EQ {dsm[j][0]}.{dsm[j][1].upper()};,$\n""")


                        elif lista_res[h][1] == 2:
                            with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:
                                arquivo.write(f"SEGMENT={dsm[j][0]}, SEGTYPE=KU,PARENT={table_name_dic[dsm[j][2]]}, CRFILE={caminho_master}/{dsm[j][0]}, CRINCLUDE=ALL , CRJOINTYPE={dsm[j][4].replace(' ', '_').upper()}, JOIN_WHERE={table_name_dic[dsm[j][2]]}.{dsm[j][3].upper()} EQ {dsm[j][0]}.{dsm[j][1].upper()} AND {table_name_dic[dsm[(j)][2]]}.{dsm[(j+1)][3].upper()} EQ {dsm[(j)][0]}.{dsm[(j+1)][1].upper()};,$\n")
                            

                        elif lista_res[h][1] == 3:
                            with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:
                                arquivo.write(f"""SEGMENT={dsm[j][0]}, SEGTYPE=KU,PARENT={table_name_dic[dsm[j][2]]}, CRFILE={caminho_master}/{dsm[j][0]}, CRINCLUDE=ALL , CRJOINTYPE={dsm[j][4].replace(' ', '_').upper()}, JOIN_WHERE={table_name_dic[dsm[j][2]]}.{dsm[j][3].upper()} EQ {dsm[j][0]}.{dsm[j][1].upper()} AND {table_name_dic[dsm[j][2]]}.{dsm[j+1][3].upper()} EQ {dsm[j][0]}.{dsm[j+1][1].upper()} AND {table_name_dic[dsm[j][2]]}.{dsm[j+2][3].upper()} EQ {dsm[j][0]}.{dsm[j+2][1].upper()};,$\n""")
                        break

                    except BaseException as err:
                        deleta_pasta(path_name)
                        msg_alerta_erro("Ocorreu um erro ao gerar as masterfiles!","Favor preencher as colunas 'DBN0 Attribute Physical Name', 'Enrichment Attribute Physical Name' e 'AdHoc Join Type' presentes na aba 3. Data Source Map.")


def mtf_prt3_sem_caminho(inventory_name,Data_Sources_Map_list,table_name_dic, path_name):                
   #Masterfiles - parte3
    dsm = Data_Sources_Map_list[:]
    
    for i in inventory_name:
        lista_ETN = []
        lista_res = []
        #Gera a lista de Enrichment Table Name
        for t in range(len(dsm)):
            if dsm[t][2] == i:
                lista_ETN.append(dsm[t][0])

        

    #Comando Set está mudando a ordem
        lista_ETN_uni = list(set(lista_ETN))


        for k in lista_ETN_uni:
            lista_res.append((k,lista_ETN.count(k)))
        
        for h in range(len(lista_res)):
            for j in range(len(dsm)):  
                
                if dsm[j][2] == i and lista_res[h][0] == dsm[j][0]:
                    if lista_res[h][1] == 1:
                        with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:
                            arquivo.write(f"""SEGMENT={dsm[j][0]}, SEGTYPE=KU,PARENT={table_name_dic[dsm[j][2]]}, CRFILE={dsm[j][0]}, CRINCLUDE=ALL , CRJOINTYPE={dsm[j][4].replace(' ', '_').upper()}, JOIN_WHERE={table_name_dic[dsm[j][2]]}.{dsm[j][3].upper()} EQ {dsm[j][0]}.{dsm[j][1].upper()};,$\n""")

                    elif lista_res[h][1] == 2:
                        with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:
                            arquivo.write(f"SEGMENT={dsm[j][0]}, SEGTYPE=KU,PARENT={table_name_dic[dsm[j][2]]}, CRFILE={dsm[j][0]}, CRINCLUDE=ALL , CRJOINTYPE={dsm[j][4].replace(' ', '_').upper()}, JOIN_WHERE={table_name_dic[dsm[j][2]]}.{dsm[j][3].upper()} EQ {dsm[j][0]}.{dsm[j][1].upper()} AND {table_name_dic[dsm[(j)][2]]}.{dsm[(j+1)][3].upper()} EQ {dsm[(j)][0]}.{dsm[(j+1)][1].upper()};,$\n")
                        

                    elif lista_res[h][1] == 3:
                        with open(f"{path_name}/{i.lower()}.mas", "a") as arquivo:
                            arquivo.write(f"""SEGMENT={dsm[j][0]}, SEGTYPE=KU,PARENT={table_name_dic[dsm[j][2]]}, CRFILE={dsm[j][0]}, CRINCLUDE=ALL , CRJOINTYPE={dsm[j][4].replace(' ', '_').upper()}, JOIN_WHERE={table_name_dic[dsm[j][2]]}.{dsm[j][3].upper()} EQ {dsm[j][0]}.{dsm[j][1].upper()} AND {table_name_dic[dsm[j][2]]}.{dsm[j+1][3].upper()} EQ {dsm[j][0]}.{dsm[j+1][1].upper()} AND {table_name_dic[dsm[j][2]]}.{dsm[j+2][3].upper()} EQ {dsm[j][0]}.{dsm[j+2][1].upper()};,$\n""")
                    break

def acx(inventory_name, Data_Sources_list, pk_list, path_name):                       
    #Criação do acx
    for i in inventory_name:
        lista_pk = []
        for ds in Data_Sources_list:
            if ds[0] == i:
                with open(f"{path_name}/{i.lower()}.acx", "a") as arquivo:
                    arquivo.write(f"SEGNAME={ds[1]},\n")
                    arquivo.write(f"    TABLENAME={ds[4].lower()}.{ds[1].lower()},\n")
                    arquivo.write(f"    CONNECTION={ds[4].upper()},\n")
                    for pk in pk_list:
                        if pk[0] == i:
                            lista_pk.append(pk[1])
                    arquivo.write(f"    KEY={'/'.join(lista_pk)}, $")
         
def cria_zip(path_name_xlsm):
    temp = path_name_xlsm.split('/')
    temp.pop()
    temp = '/'.join(temp)
    path_name = temp + '/MASTERFILES'

    if not os.path.exists(path_name+',zip'):
        z = zipfile.ZipFile(f'{temp}/MASTERFILES.zip', 'w', zipfile.ZIP_DEFLATED)
        z.write(path_name)
        z.close()